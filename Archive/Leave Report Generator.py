# Core Python system interface - Used for system-specific parameters and functions
# Security: Relatively safe, but sys.exit() could be abused to terminate programs unexpectedly
import sys

# Spawns new processes, runs external commands
# Security: Can execute arbitrary system commands. But required System Administrator to execute commands.
import subprocess

# Operating system interface - file/directory operations
# Security: Ensure program can only access the folder it is in, and not directly the network or C drive. Check permissions
import os

# GUI framework for desktop applications
# Security: Safe for desktop GUI creation
import tkinter as tk

# Pop-up message boxes for tkinter
# Security: Safe, but don't display unsanitized user input
from tkinter import messagebox

# Date and time handling utilities
# Security: Safe to use, no major security concerns
from datetime import datetime, timedelta

# Data manipulation and analysis library
# Security: Generally safe, but be careful with untrusted CSV/Excel files due to formula injection
import pandas as pd

# Time zone handling utilities
# Security: Safe to use, no major security concerns
import pytz

# Excel file creation and manipulation
# Security: Sanitize inputs before writing to cells
from openpyxl import Workbook

# Excel styling capabilities
# Security: Safe to use, no major security concerns
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Excel column utilities
# Security: Safe to use, no major security concerns
from openpyxl.utils import get_column_letter

# Time-related functions and utilities
# Security: Safe to use, but time.sleep() could be used for DOS attacks if not limited
import time

# Popup Message for displaying custom user warning.
# Security:Just a opup box. 
import warnings


# Suppress the specific warning about header/footer parsing
warnings.filterwarnings("ignore", message="Cannot parse header or footer so it will be ignored")

#Combined two tabs, as through they are the same areas
COMBINED_LOCATIONS = [
    ("ADELAIDE HILLS & STRATHALBYN", ["ADELAIDE HILLS", "STRATHALBYN"]),
    # Add more combinations as needed, for example:
    # ("North & South", ["North", "South"]),
    # ("East & West", ["East", "West"]),
]


DATE_COLUMN_WIDTH = 10.00  # Width in Excel units for date columns
CREATE_INDIVIDUAL_TABS_FOR_COMBINED_LOCATIONS = False  # Set to True if you want individual tabs for combined locations

def show_error_popup(message: str):
    # Create the main tkinter window
    root = tk.Tk()
    # Hide the main window - we only want to show the message box
    root.withdraw()
    # Display error message in a popup
    messagebox.showerror("Error", message)
    # Clean up by destroying the root window
    root.destroy()  # Fixed typo in 'rot.destroy()'

def find_excel_files(directory: str):
    # Check if the directory exists
    if not os.path.exists(directory):
        show_error_popup(f"Error: The 'Reports' folder does not exist at {directory}")
        return None, None
    
    # Get list of Excel files in directory using list comprehension
    # Filters for .xlsx and .xls extensions
    excel_files = [f for f in os.listdir(directory) if f.endswith(('.xlsx', '.xls'))]
    
    # Validate that exactly 2 Excel files were found
    if len(excel_files) != 2:
        show_error_popup(f"Error: Expected 2 Excel files, found {len(excel_files)}")
        return None, None
    
    # Return tuple of full file paths using os.path.join for safe path concatenation
    return tuple(os.path.join(directory, f) for f in excel_files)

def load_excel_file(file_path: str):
    try:
        # Select appropriate engine based on file extension
        engine = 'openpyxl' if file_path.endswith('.xlsx') else 'xlrd'
        # Load the Excel file into a pandas DataFrame
        return pd.read_excel(file_path, engine=engine)
    except PermissionError:
        # Handle case where file is locked by another process
        show_error_popup(f"Error: Unable to open '{os.path.basename(file_path)}'. The file may be open in another program.")
    except Exception as e:
        # Catch all other exceptions and show generic error message
        show_error_popup(f"Error reading {os.path.basename(file_path)}: {str(e)}")
    return None

def validate_headers(df: pd.DataFrame, expected_headers: list):
    # Check if all expected headers exist in DataFrame columns
    return all(header in df.columns for header in expected_headers)

def load_and_validate_data(file_path: str, leave_headers: list, work_areas_headers: list):    
    # Attempt to load the Excel file
    df = load_excel_file(file_path)
    
    if df is not None:
        # Check if file matches Employee Leave format
        if validate_headers(df, leave_headers):
            print(f"Successfully loaded and validated {os.path.basename(file_path)} as Employee Leave")
            return df, "Employee Leave"
        # Check if file matches Employee Work Areas format
        elif validate_headers(df, work_areas_headers):
            print(f"Successfully loaded and validated {os.path.basename(file_path)} as Employee Work Areas")
            return df, "Employee Work Areas"
        else:
            # Neither format matches - show error
            show_error_popup(f"Error: {os.path.basename(file_path)} does not have the expected headers for either dataset")
    return None, ""

def process_leave_requests(df: pd.DataFrame):
    df['Employee_Code'] = df['Employee_Code'].astype(str).str.lstrip('E')
    df['Start_Time'] = pd.to_datetime(df['Start_Time'])
    df['End_Time'] = pd.to_datetime(df['End_Time'])

    expanded_leaves = []
    for _, row in df.iterrows():
        start_date = row['Start_Time'].date()
        end_date = row['End_Time'].date()
        
        if row['End_Time'].time() == pd.Timestamp('00:00:00').time():
            end_date -= timedelta(days=1)
        
        for date in pd.date_range(start=start_date, end=end_date):
            expanded_leaves.append({
                'Employee_Code': row['Employee_Code'],
                'Employee_Name': row['Employee_Name'],
                'Shift_Type': row['Shift_Type'],
                'Leave_Date': date,
                'Status': row['Status']
            })

    expanded_df = pd.DataFrame(expanded_leaves)
    return expanded_df.groupby(['Employee_Code', 'Leave_Date']).first().reset_index()

def create_summary_tables(employee_data: pd.DataFrame, leave_requests: pd.DataFrame):
    employees_with_leave = leave_requests['Employee_Code'].unique()
    active_employees = employee_data[employee_data['Employee_Code'].isin(employees_with_leave)]

    return {
        'employee_types': employee_data['Employment_Type_Name'].unique().tolist(),
        'statuses': leave_requests['Status'].unique().tolist(),
        'shift_types': leave_requests['Shift_Type'].unique().tolist(),
        'active_employees': active_employees
    }

def process_employee_work_areas(df: pd.DataFrame):
    employee_data = df[['Employee_Code', 'Employee_Name', 'Employment_Type_Name', 'Location', 'Department']].drop_duplicates()
    employee_data['Employee_Code'] = employee_data['Employee_Code'].astype(str).str.lstrip('E')
    return employee_data

def process_data():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    downloads_dir = os.path.join(script_dir, 'Reports')

    leave_headers = ['Employee_Code', 'Employee_Name', 'Shift_Type', 'Start_Time', 'End_Time', 'Status']
    work_areas_headers = ['Employee_Code', 'Employee_Name', 'Employment_Type_Name', 'Location', 'Department', 'Role']

    file1, file2 = find_excel_files(downloads_dir)
    if not file1 or not file2:
        return None, None, None

    df1, type1 = load_and_validate_data(file1, leave_headers, work_areas_headers)
    df2, type2 = load_and_validate_data(file2, leave_headers, work_areas_headers)

    if df1 is None or df2 is None:
        return None, None, None

    if type1 == "Employee Leave" and type2 == "Employee Work Areas":
        leave_requests, employee_work_areas = df1, df2
    elif type1 == "Employee Work Areas" and type2 == "Employee Leave":
        leave_requests, employee_work_areas = df2, df1
    else:
        show_error_popup("Error: Unable to determine which file is which")
        return None, None, None

    print("Data loaded successfully")

    employee_data = process_employee_work_areas(employee_work_areas)
    processed_leave_requests = process_leave_requests(leave_requests)

    processed_leave_requests = pd.merge(processed_leave_requests,
                                        employee_data[['Employee_Code', 'Department', 'Location']],
                                        on='Employee_Code',
                                        how='left')

    summary_data = create_summary_tables(employee_data, processed_leave_requests)

    return employee_data, processed_leave_requests, summary_data

def create_color_mapping(statuses):
    colors = ['FF0000', '00FF00', '0000FF', 'FFFF00', '00FFFF', 'FF00FF']
    return {status: PatternFill(start_color=color, end_color=color, fill_type="solid")
            for status, color in zip(statuses, colors[:len(statuses)])}

def create_status_abbreviations(statuses):
    abbr_dict = {}
    for status in statuses:
        abbr = status[0].upper()
        counter = 1
        while abbr in abbr_dict.values():
            abbr = status[0].upper() + str(counter)
            counter += 1
        abbr_dict[status] = abbr
    return abbr_dict

def apply_table_styling(ws, start_row, end_row, start_col, end_col, thick_border=False):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                          bottom=Side(style='medium'))
    
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if thick_border:
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            row[0].border = Border(left=Side(style='medium'), top=row[0].border.top, bottom=row[0].border.bottom)
            row[-1].border = Border(right=Side(style='medium'), top=row[-1].border.top, bottom=row[-1].border.bottom)
        for col in ws.iter_cols(min_col=start_col, max_col=end_col, min_row=start_row, max_row=end_row):
            col[0].border = Border(top=Side(style='medium'), left=col[0].border.left, right=col[0].border.right)
            col[-1].border = Border(bottom=Side(style='medium'), left=col[-1].border.left, right=col[-1].border.right)

def format_date(date):
    return f"{date.strftime('%a')} {date.strftime('%d').lstrip('0')}/{date.strftime('%m').lstrip('0')}"

def create_employee_table(ws, employee_data, leave_requests, start_date, end_date, color_mapping, status_abbr, row_offset=0):
    headers = ["Employee Name", "Type", "Days"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_offset+1, column=col, value=header)
        cell.font = Font(bold=True)

    date_col_start = len(headers) + 1
    for col, date in enumerate(pd.date_range(start=start_date, end=end_date), start=date_col_start):
        cell = ws.cell(row=row_offset+1, column=col, value=format_date(date))
        cell.font = Font(bold=True)
        if is_weekend(date):
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # Ensure employee_data has unique entries
    employee_data = employee_data.drop_duplicates(subset=['Employee_Code'])

    SHOW_ALL_EMPLOYEES = False
    if not SHOW_ALL_EMPLOYEES:
        employees_with_leave = leave_requests['Employee_Code'].unique()
        employee_data = employee_data[employee_data['Employee_Code'].isin(employees_with_leave)]

    # Sort employee_data by Employment Type and then by Employee Name
    employee_data = employee_data.sort_values(['Employment_Type_Name', 'Employee_Name'])

    row = row_offset + 1

    if not employee_data.empty:
        for row, employee in enumerate(employee_data.itertuples(), start=row_offset+2):
            ws.cell(row=row, column=1, value=employee.Employee_Name)
            ws.cell(row=row, column=2, value=employee.Employment_Type_Name)

            employee_leaves = leave_requests[leave_requests['Employee_Code'] == employee.Employee_Code]
            days_requested = employee_leaves['Leave_Date'].nunique()
            ws.cell(row=row, column=3, value=days_requested)

            for col, date in enumerate(pd.date_range(start=start_date, end=end_date), start=date_col_start):
                leave_on_date = employee_leaves[employee_leaves['Leave_Date'] == date]
                cell = ws.cell(row=row, column=col)
                if not leave_on_date.empty:
                    status = leave_on_date.iloc[0]['Status']
                    cell.value = status_abbr[status]
                    cell.fill = color_mapping[status]
                elif is_weekend(date):
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    else:
        row += 1
        ws.cell(row=row, column=1, value="No employees to display")

    apply_table_styling(ws, row_offset+1, row, 1, date_col_start + (end_date - start_date).days, thick_border=True)

    return row, date_col_start

def create_department_table(ws, leave_requests, departments, start_date, end_date, employee_data, row_offset=0, date_col_start=0):
    ws.cell(row=row_offset + 1, column=1, value="Department").font = Font(bold=True)
    ws.cell(row=row_offset + 1, column=2, value="Count").font = Font(bold=True)

    for col, date in enumerate(pd.date_range(start=start_date, end=end_date), start=date_col_start):
        cell = ws.cell(row=row_offset + 1, column=col, value=format_date(date))
        cell.font = Font(bold=True)
        if is_weekend(date):
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    for row, department in enumerate(departments, start=row_offset + 2):
        ws.cell(row=row, column=1, value=department)
        dept_employees = employee_data[employee_data['Department'] == department]
        total_employees = len(dept_employees)
        ws.cell(row=row, column=2, value=total_employees)

        dept_leaves = leave_requests[leave_requests['Department'] == department]

        for col, date in enumerate(pd.date_range(start=start_date, end=end_date), start=date_col_start):
            leave_on_date = dept_leaves[dept_leaves['Leave_Date'] == date]
            cell = ws.cell(row=row, column=col)
            if not leave_on_date.empty:
                total_on_leave = leave_on_date['Employee_Code'].nunique()
                cell.value = total_on_leave
            if is_weekend(date):
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    apply_table_styling(ws, row_offset+1, row, 1, date_col_start + (end_date - start_date).days, thick_border=True)

    return row

def create_status_summary_table(ws, statuses, start_date, end_date, color_mapping, status_abbr, row_offset=0, date_col_start=0):
    for row, status in enumerate(statuses, start=row_offset + 1):
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=2, value=status_abbr[status])
        ws.cell(row=row, column=2).fill = color_mapping[status]

        for col in range(date_col_start, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            data_range = f'${ws.cell(row=2, column=col).column_letter}$2:${ws.cell(row=2, column=col).column_letter}${row_offset}'
            cell.value = f'=COUNTIF({data_range},"{status_abbr[status]}")'
            if is_weekend(start_date + timedelta(days=col - date_col_start)):
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    total_row = row + 1
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)

    for col in range(date_col_start, ws.max_column + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.value = f'=SUM({ws.cell(row=row_offset + 1, column=col).coordinate}:{ws.cell(row=total_row - 1, column=col).coordinate})'
        if is_weekend(start_date + timedelta(days=col - date_col_start)):
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    apply_table_styling(ws, row_offset + 1, total_row, 1, ws.max_column, thick_border=True)

    return total_row

def create_comprehensive_leave_report(employee_data, leave_requests, statuses, locations):
    color_mapping = create_color_mapping(statuses)
    status_abbr = create_status_abbreviations(statuses)

    wb = Workbook()

    sheets = {'Global': wb.active}
    sheets['Global'].title = "Global"

    # Create a set of locations that are part of combined locations
    combined_locations_set = set()
    for _, component_locations in COMBINED_LOCATIONS:
        combined_locations_set.update(component_locations)

    # Create sheets based on the configuration
    for location in locations:
        if location != 'Global':
            if location not in combined_locations_set or CREATE_INDIVIDUAL_TABS_FOR_COMBINED_LOCATIONS:
                sheets[location] = wb.create_sheet(title=location)

    # Add combined location sheets
    for combined_name, _ in COMBINED_LOCATIONS:
        sheets[combined_name] = wb.create_sheet(title=combined_name)

    # Use the full date range from min to max date in leave requests
    start_date = leave_requests['Leave_Date'].min()
    end_date = leave_requests['Leave_Date'].max()
    date_range = pd.date_range(start=start_date, end=end_date)

    for sheet_name, ws in sheets.items():
        if sheet_name == 'Global':
            active_employees = employee_data
            leave_requests_filtered = leave_requests
        elif sheet_name in dict(COMBINED_LOCATIONS):
            component_locations = dict(COMBINED_LOCATIONS)[sheet_name]
            active_employees = employee_data[employee_data['Location'].isin(component_locations)]
            leave_requests_filtered = leave_requests[leave_requests['Location'].isin(component_locations)]
        else:
            active_employees = employee_data[employee_data['Location'] == sheet_name]
            leave_requests_filtered = leave_requests[leave_requests['Location'] == sheet_name]

        departments = active_employees['Department'].unique()

        last_employee_row, date_col_start = create_employee_table(ws, active_employees, leave_requests_filtered, start_date, end_date, color_mapping, status_abbr, row_offset=0)
        last_summary_row = create_status_summary_table(ws, statuses, start_date, end_date, color_mapping, status_abbr, row_offset=last_employee_row, date_col_start=date_col_start)
        last_department_row = create_department_table(ws, leave_requests_filtered, departments, start_date, end_date, active_employees, row_offset=last_summary_row + 1, date_col_start=date_col_start)
        apply_styling_to_worksheet(ws, start_date, end_date, color_mapping)

    return wb

def is_weekend(date, timezone='Australia/Sydney'):
    tz = pytz.timezone(timezone)
    date_in_tz = tz.localize(datetime(date.year, date.month, date.day))
    return date_in_tz.weekday() >= 5 

def apply_styling_to_worksheet(ws, start_date, end_date, color_mapping):
    ws.freeze_panes = 'C2'
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical='center')
            
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif cell.row == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Adjust column widths
    for col in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col)
        if col <= 3:  # First three columns
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        else:  # Date columns
            ws.column_dimensions[column_letter].width = DATE_COLUMN_WIDTH

    print(f"Creating Worksheet: {ws.title}")

def save_workbook_safely(workbook, filename):
    try:
        workbook.save(filename)        
        time.sleep(10)
        print(f"Successfully saved.")
        return True
    except PermissionError:
        show_error_popup(f"Error: Unable to save '{filename}'. The file may be open in another program. "
                         "Please close the file and run the script again.")
        return False
    except Exception as e:
        show_error_popup(f"Error: An unexpected error occurred while saving '{filename}': {str(e)}")
        return False

def keep_console_open():
    print("\nPress Enter to exit this program...")
    input()

def delete_old_report(filename):
    try:
        if os.path.exists(filename):
            os.remove(filename)
            time.sleep(5)
            print(f"Successfully deleted old report: {filename}")
        return True
    except PermissionError:
        show_error_popup(f"Error: Unable to delete '{filename}'. The file may be open in another program / user. "
                         "Please close the file and run the script again.")
        return False
    except Exception as e:
        show_error_popup(f"Error: An unexpected error occurred while deleting '{filename}': {str(e)}")
        return False

def main():
    try:        
        print("Processing data...")
        employee_data, processed_leave_requests, summary_data = process_data()

        if employee_data is None or processed_leave_requests is None or summary_data is None:
            print("Failed to generate report due to data processing errors.")
            return

        print("Setting up report...")
        statuses = summary_data['statuses']
        locations = employee_data['Location'].unique().tolist()
        if 'Global' not in locations:
            locations = ['Global'] + locations

        output_filename = 'Comprehensive Leave Report.xlsx'
        
        print(f"Attempting to delete old report: {output_filename}")
        if not delete_old_report(output_filename):
            print("Failed to delete the old report. Please check the error message and try again.")
            return

        print("Generating report...")
        comprehensive_report = create_comprehensive_leave_report(employee_data, processed_leave_requests, statuses,
                                                                 locations)
        print("Saving report...")
        if save_workbook_safely(comprehensive_report, output_filename):
            print(f"Saved as '{output_filename}'")
        else:
            print("Failed to save the report. Please check the error message and try again.")
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}")
    finally:
        keep_console_open()

if __name__ == "__main__":
    main()
