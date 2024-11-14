import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class LeaveReportGenerator:
    def __init__(self, employees, employee_manager):
        self.employees = {code: emp for code, emp in employees.items() if len(emp.leave_dates) > 0}
        self.employee_manager = employee_manager  # Store the complete employee_manager
        self.wb = Workbook()
        self.ws = None
        self.status_colors = employee_manager.leave_status_manager.status_colors
        self.COMBINED_LOCATIONS = [("ADELAIDE HILLS & STRATHALBYN", ["ADELAIDE HILLS", "STRATHALBYN"]), ]

        # Define border styles
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))

        self.thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                                   bottom=Side(style='medium'))

    def _get_all_leave_dates(self):
        """Collect all unique leave dates from all employees."""
        all_dates = set()
        for employee in self.employees.values():
            all_dates.update(employee.leave_dates.keys())
        return sorted(list(all_dates))

    def _format_date_header(self, date):
        """Format date as 'Mon 25/11'."""
        return date.strftime('%a %d/%m')

    def _get_status_initial(self, status):
        """Get first character of status."""
        return status[0] if status else ''

    def _get_all_locations(self):
        """Get all unique locations including combined ones."""
        locations = set()
        for employee in self.employees.values():
            for work_area in employee.work_areas:
                locations.add(work_area.location)
        return sorted(list(locations))

    def _get_employees_for_location(self, location, include_combined=False):
        """Get employees for a specific location or combined locations."""
        if isinstance(location, tuple):
            # Handle combined locations
            combined_name, location_list = location
            return {code: emp for code, emp in self.employees.items() if
                    any(area.location in location_list for area in emp.work_areas)}
        else:
            # Handle single location
            return {code: emp for code, emp in self.employees.items() if
                    any(area.location == location for area in emp.work_areas)}

    def _calculate_status_totals(self, filtered_employees, all_dates):
        """Calculate totals per status and overall total for each date."""
        status_totals = {}
        overall_totals = {}

        # Initialize totals dictionaries
        for date in all_dates:
            status_totals[date] = {}
            overall_totals[date] = 0

        # Calculate totals
        for employee in filtered_employees.values():
            for date in all_dates:
                if date in employee.leave_dates:
                    status = employee.leave_dates[date].status
                    if status not in status_totals[date]:
                        status_totals[date][status] = 0
                    status_totals[date][status] += 1
                    overall_totals[date] += 1

        return status_totals, overall_totals

    def _generate_worksheet(self, ws_name, filtered_employees, all_dates):
        """Generate a worksheet for the given employees."""
        ws = self.wb.create_sheet(ws_name)

        # Get location from worksheet name (keep existing location logic)
        location = ws_name
        if location == "GLOBAL":
            location = None
        elif any(combined[0][:31] == location for combined in self.COMBINED_LOCATIONS):
            location = next(combined for combined in self.COMBINED_LOCATIONS if combined[0][:31] == location)

        # Set up headers
        headers = ['Employee Name (Code)', 'Employment Type', 'Leave Count']
        headers.extend([self._format_date_header(date) for date in all_dates])
        max_col = len(headers)

        # Write headers with proper borders
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

            # Apply border - thicker on top and appropriate sides
            border_style = {
                'left': Side(style='medium' if col == 1 else 'thin'),
                'right': Side(style='medium' if col == max_col else 'thin'),
                'top': Side(style='medium'),
                'bottom': Side(style='thin')
            }
            cell.border = Border(**border_style)

        # Write employee data
        row = 2
        last_data_row = row + len(filtered_employees) - 1  # Calculate last row for border handling

        for emp_code, employee in sorted(filtered_employees.items(), key=lambda x: x[1].name):
            for col in range(1, max_col + 1):
                # Initialize all cells with basic border
                cell = ws.cell(row=row, column=col)
                border_style = {
                    'left': Side(style='medium' if col == 1 else 'thin'),
                    'right': Side(style='medium' if col == max_col else 'thin'),
                    'top': Side(style='thin'),
                    'bottom': Side(style='medium' if row == last_data_row else 'thin')
                }
                cell.border = Border(**border_style)
                cell.alignment = Alignment(horizontal='center')

            # Employee name and code
            cell = ws.cell(row=row, column=1)
            cell.value = f"{employee.name} ({emp_code})"
            cell.alignment = Alignment(horizontal='left')  # Left align names

            # Employment type
            cell = ws.cell(row=row, column=2)
            cell.value = employee.employment_type

            # Leave count
            cell = ws.cell(row=row, column=3)
            cell.value = len(employee.leave_dates)

            # Status for each date
            for col, date in enumerate(all_dates, 4):
                cell = ws.cell(row=row, column=col)
                if date in employee.leave_dates:
                    leave_date = employee.leave_dates[date]
                    status = self._get_status_initial(leave_date.status)
                    cell.value = status
                    cell.fill = PatternFill(start_color=self.status_colors[leave_date.status],
                                            end_color=self.status_colors[leave_date.status],
                                            fill_type='solid')

            row += 1

        row += 2  # Add spacing between tables

        # Generate department table
        self._add_department_leave_table(ws, row, filtered_employees, all_dates, location)

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            if col <= 2:
                ws.column_dimensions[column_letter].width = 30
            elif col == 3:
                ws.column_dimensions[column_letter].width = 8
            else:
                ws.column_dimensions[column_letter].width = 12

    def _get_color_for_ratio(self, on_leave, total):
        """Calculate cell color based on ratio of employees on leave."""
        if total == 0 or on_leave == 0:
            return None
        ratio = on_leave / total
        # Convert ratio to hex color from white (FFFFFF) to red (FF0000)
        intensity = int(ratio * 255)
        red = 255
        green = blue = 255 - intensity
        return f"{red:02x}{green:02x}{blue:02x}"

    def _get_all_employees_by_department(self, employee_manager):
        """Get unique employee count per department including ALL employees."""
        department_counts = {}
        # Use the complete employee set from employee_manager, not just filtered_employees
        for emp_code, emp in employee_manager.employees.items():
            for work_area in emp.work_areas:
                dept = work_area.department
                if dept not in department_counts:
                    department_counts[dept] = set()
                department_counts[dept].add(emp_code)
        return {dept: len(employees) for dept, employees in department_counts.items()}

    def _add_department_leave_table(self, ws, start_row, filtered_employees, all_dates, location):
        row = start_row

        # Get departments and counts filtered by location
        department_counts = self._get_department_employee_counts(location, self.employee_manager)
        departments = self._get_departments_for_location(location, self.employee_manager)

        # Table headers
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1)
        cell.value = 'Department'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = self.thin_border

        cell = ws.cell(row=row, column=3)
        cell.value = 'Employees'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = self.thin_border

        # Date headers
        for col, date in enumerate(all_dates, 4):
            cell = ws.cell(row=row, column=col)
            cell.value = self._format_date_header(date)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thin_border

        last_data_row = row + len(departments)  # Calculate last row for border handling
        row += 1

        for dept in departments:
            # Department name
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            cell = ws.cell(row=row, column=1)
            cell.value = dept
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='left')

            # Employee count
            total_employees = department_counts[dept]
            cell = ws.cell(row=row, column=3)
            cell.value = total_employees
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center')

            # Daily counts with color scaling
            for col, date in enumerate(all_dates, 4):
                cell = ws.cell(row=row, column=col)
                on_leave = sum(1 for emp in filtered_employees.values() if
                               date in emp.leave_dates and
                               any(wa.department == dept for wa in emp.work_areas))

                if on_leave > 0:
                    cell.value = on_leave
                    color = self._get_color_for_ratio(on_leave, total_employees)
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center')

            row += 1

        # Handle borders for merged cells and table outline
        max_col = len(all_dates) + 3
        for row_num in range(start_row, last_data_row + 1):
            # Merged cells (columns 1 and 2)
            ws.cell(row=row_num, column=1).border = Border(
                left=Side(style='medium'),
                right=Side(style='thin'),
                top=Side(style='medium') if row_num == start_row else Side(style='thin'),
                bottom=Side(style='medium') if row_num == last_data_row else Side(style='thin')
            )
            ws.cell(row=row_num, column=2).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='medium') if row_num == start_row else Side(style='thin'),
                bottom=Side(style='medium') if row_num == last_data_row else Side(style='thin')
            )

            # Regular cells
            for col in range(3, max_col + 1):
                border_style = {
                    'left': Side(style='thin'),
                    'right': Side(style='medium') if col == max_col else Side(style='thin'),
                    'top': Side(style='medium') if row_num == start_row else Side(style='thin'),
                    'bottom': Side(style='medium') if row_num == last_data_row else Side(style='thin')
                }
                ws.cell(row=row_num, column=col).border = Border(**border_style)

        return row + 2

    def generate_report(self):
        """Generate the leave report in Excel format."""
        # Get all unique dates
        all_dates = self._get_all_leave_dates()

        # Remove default sheet if it exists
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

        # Generate GLOBAL worksheet
        self._generate_worksheet('GLOBAL', self.employees, all_dates)

        # Generate location-specific worksheets
        locations = self._get_all_locations()
        for location in locations:
            filtered_employees = self._get_employees_for_location(location)
            if filtered_employees:  # Only create sheet if there are employees
                safe_name = str(location)[:31]  # Excel worksheet names limited to 31 chars
                self._generate_worksheet(safe_name, filtered_employees, all_dates)

        # Generate combined location worksheets
        for combined_name, location_list in self.COMBINED_LOCATIONS:
            filtered_employees = self._get_employees_for_location((combined_name, location_list))
            if filtered_employees:
                safe_name = combined_name[:31]
                self._generate_worksheet(safe_name, filtered_employees, all_dates)

    def save_report(self, directory):
        """Save the report with the specified naming convention."""
        current_date = datetime.now()
        month_year = current_date.strftime("%d %b %Y")
        filename = f"Leave Report {month_year}.xlsx"
        filepath = os.path.join(directory, filename)
        self.wb.save(filepath)
        print(f"Saved: {os.path.basename(filename)}")

    def _get_departments_for_location(self, location, employee_manager):
        """Get departments for a specific location or combined locations."""
        departments = set()
        if isinstance(location, tuple):
            # Handle combined locations
            combined_name, location_list = location
            for emp in employee_manager.employees.values():
                if any(area.location in location_list for area in emp.work_areas):
                    departments.update(area.department for area in emp.work_areas
                                       if area.location in location_list)
        else:
            # Handle single location
            for emp in employee_manager.employees.values():
                if any(area.location == location for area in emp.work_areas):
                    departments.update(area.department for area in emp.work_areas
                                       if area.location == location)
        return sorted(list(departments))

    def _get_department_employee_counts(self, location, employee_manager):
        """Get unique employee count per department for a specific location."""
        department_counts = {}

        if isinstance(location, tuple):
            # Handle combined locations
            combined_name, location_list = location
            for emp_code, emp in employee_manager.employees.items():
                if any(area.location in location_list for area in emp.work_areas):
                    for work_area in emp.work_areas:
                        if work_area.location in location_list:
                            dept = work_area.department
                            if dept not in department_counts:
                                department_counts[dept] = set()
                            department_counts[dept].add(emp_code)
        else:
            # Handle single location
            for emp_code, emp in employee_manager.employees.items():
                if any(area.location == location for area in emp.work_areas):
                    for work_area in emp.work_areas:
                        if work_area.location == location:
                            dept = work_area.department
                            if dept not in department_counts:
                                department_counts[dept] = set()
                            department_counts[dept].add(emp_code)

        return {dept: len(employees) for dept, employees in department_counts.items()}


def generate_leave_report(employees, output_directory, employee_manager):
    """Main function to generate and save the leave report."""
    report_generator = LeaveReportGenerator(employees, employee_manager)
    report_generator.generate_report()
    report_generator.save_report(output_directory)
