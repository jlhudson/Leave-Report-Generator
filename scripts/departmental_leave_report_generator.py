import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class DepartmentalLeaveReportGenerator:
    def __init__(self, employees, employee_manager):
        self.employees = {code: emp for code, emp in employees.items() if len(emp.leave_dates) > 0}
        self.wb = Workbook()
        self.ws = None
        self.status_colors = employee_manager.leave_status_manager.status_colors
        self.COMBINED_LOCATIONS = [("ADELAIDE HILLS & STRATHALBYN", ["ADELAIDE HILLS", "STRATHALBYN"]), ]

        # Define styles
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))

        self.thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                                   bottom=Side(style='medium'))

        self.header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        self.total_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')

    def _generate_worksheet(self, ws_name, location_employees, all_dates):
        """Generate a worksheet for the given location."""
        ws = self.wb.create_sheet(ws_name)

        # Write date headers only once at the top
        current_row = 1

        # Empty cell in top-left corner
        cell = ws.cell(row=current_row, column=1)
        cell.fill = self.header_fill
        cell.border = self.thick_border

        # Write date headers
        for col, date in enumerate(all_dates, start=2):
            cell = ws.cell(row=current_row, column=col)
            cell.value = self._format_date_header(date)
            cell.font = Font(bold=True)
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thick_border

        # Move to next row for department data
        current_row += 1

        # Get departments and their employees
        dept_employees = self._get_location_departments(location_employees)

        # Create a thicker border for department tables
        department_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        # Create a border with only the right side for internal cells
        right_border = Border(right=Side(style='thin'))

        # Process each department
        first_department = True
        for dept_name, dept_emps in dept_employees.items():
            if not first_department:
                current_row += 1  # Add more spacing between departments
            first_department = False

            max_employees_per_day = 0
            employees_by_date = {}

            # Pre-calculate employees for each date and find maximum
            for date in all_dates:
                emps_on_leave = self._get_employees_by_date(dept_emps, date)
                employees_by_date[date] = emps_on_leave
                max_employees_per_day = max(max_employees_per_day, len(emps_on_leave))

            # Write department name only if there are employees
            if max_employees_per_day > 0:
                # Store starting positions for the department table
                table_start_row = current_row
                table_start_col = 1
                table_end_col = len(all_dates) + 1

                # Write department name
                cell = ws.cell(row=current_row, column=1)
                cell.value = dept_name
                cell.font = Font(bold=True)
                cell.fill = self.header_fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                # Write employee names
                start_row = current_row  # Start merging from the first employee row
                for row_offset in range(max_employees_per_day):
                    for col, date in enumerate(all_dates, start=2):
                        cell = ws.cell(row=current_row + row_offset, column=col)

                        emps_on_leave = employees_by_date[date]
                        if row_offset < len(emps_on_leave):
                            emp, status = emps_on_leave[row_offset]
                            cell.value = self._format_employee_name(emp.name)
                            cell.fill = PatternFill(start_color=self.status_colors[status],
                                                    end_color=self.status_colors[status], fill_type='solid')
                            cell.alignment = Alignment(horizontal='center')

                        # Add right border for all cells except the last column
                        if col < table_end_col:
                            cell.border = right_border

                current_row += max_employees_per_day

                # Add totals row (without borders)
                total_row = current_row

                # Write empty cell with total formatting
                cell = ws.cell(row=total_row, column=1)
                cell.fill = self.total_fill

                # Calculate and write totals for each date
                for col, date in enumerate(all_dates, start=2):
                    count = self._count_unique_employees_on_date(dept_emps, date)
                    cell = ws.cell(row=total_row, column=col)
                    cell.value = count if count > 0 else ''
                    cell.font = Font(bold=True)
                    cell.fill = self.total_fill
                    cell.alignment = Alignment(horizontal='center')
                    # Add right border for all cells except the last column
                    if col < table_end_col:
                        cell.border = right_border

                # Merge department name cells
                ws.merge_cells(start_row=start_row, start_column=1,
                               end_row=total_row, end_column=1)

                # Apply thick border around the entire department table
                for row in range(table_start_row, total_row + 1):
                    # Left border of first column
                    ws.cell(row=row, column=table_start_col).border = Border(
                        left=Side(style='medium'),
                        right=Side(style='thin'),
                        top=Side(style='medium') if row == table_start_row else None,
                        bottom=Side(style='medium') if row == total_row else None
                    )
                    # Right border of last column
                    ws.cell(row=row, column=table_end_col).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='medium'),
                        top=Side(style='medium') if row == table_start_row else None,
                        bottom=Side(style='medium') if row == total_row else None
                    )
                    # Top border for first row
                    if row == table_start_row:
                        for col in range(table_start_col + 1, table_end_col):
                            ws.cell(row=row, column=col).border = Border(
                                top=Side(style='medium'),
                                right=Side(style='thin')
                            )
                    # Bottom border for last row
                    elif row == total_row:
                        for col in range(table_start_col + 1, table_end_col):
                            ws.cell(row=row, column=col).border = Border(
                                bottom=Side(style='medium'),
                                right=Side(style='thin')
                            )

                current_row += 1  # Move to next row after totals

        # Adjust column widths
        for col in range(1, len(all_dates) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Freeze panes - freeze top row and first column
        ws.freeze_panes = 'B2'

    def generate_report(self):
        """Generate the departmental leave report in Excel format."""
        # Get all unique dates
        all_dates = self._get_all_leave_dates()

        # Remove default sheet if it exists
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

        # Generate GLOBAL worksheet first
        self._generate_global_worksheet(all_dates)

        # Generate location-specific worksheets
        locations = self._get_all_locations()
        for location in locations:
            filtered_employees = self._get_employees_for_location(location)
            if filtered_employees:  # Only create sheet if there are employees
                safe_name = str(location)[:31]  # Excel worksheet names limited to 31 chars
                self._generate_worksheet(safe_name, filtered_employees, all_dates)

        # Generate combined location worksheets
        for combined_name, location_list in self.COMBINED_LOCATIONS:
            filtered_employees = self._get_employees_for_location((combined_name, location_list))
            if filtered_employees:
                safe_name = combined_name[:31]
                self._generate_worksheet(safe_name, filtered_employees, all_dates)

    def _generate_global_worksheet(self, all_dates):
        """Generate a global worksheet showing all locations."""
        ws = self.wb.create_sheet("GLOBAL")

        # Write date headers only once at the top
        current_row = 1

        # Empty cell in top-left corner
        cell = ws.cell(row=current_row, column=1)
        cell.fill = self.header_fill
        cell.border = self.thick_border

        # Write date headers
        for col, date in enumerate(all_dates, start=2):
            cell = ws.cell(row=current_row, column=col)
            cell.value = self._format_date_header(date)
            cell.font = Font(bold=True)
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thick_border

        # Move to next row for location data
        current_row += 1

        locations = self._get_all_locations()
        first_location = True

        # Create a border with only the right side for internal cells
        right_border = Border(right=Side(style='thin'))

        for location in locations:
            if not first_location:
                current_row += 1  # Add more spacing between locations
            first_location = False

            filtered_employees = self._get_employees_for_location(location)
            if not filtered_employees:
                continue

            # Pre-calculate employees for each date and find maximum
            employees_by_date = {}
            max_employees_per_day = 0

            sorted_employees = sorted(filtered_employees.values(), key=lambda x: (-len(x.leave_dates), x.name))
            for date in all_dates:
                emps_on_leave = self._get_employees_by_date(sorted_employees, date)
                employees_by_date[date] = emps_on_leave
                max_employees_per_day = max(max_employees_per_day, len(emps_on_leave))

            if max_employees_per_day > 0:
                # Store starting positions for the location table
                table_start_row = current_row
                table_start_col = 1
                table_end_col = len(all_dates) + 1

                # Write location name
                cell = ws.cell(row=current_row, column=1)
                cell.value = location
                cell.font = Font(bold=True)
                cell.fill = self.header_fill
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                start_row = current_row  # Start merging from the first employee row

                # Write employee names
                for row_offset in range(max_employees_per_day):
                    for col, date in enumerate(all_dates, start=2):
                        cell = ws.cell(row=current_row + row_offset, column=col)

                        emps_on_leave = employees_by_date[date]
                        if row_offset < len(emps_on_leave):
                            emp, status = emps_on_leave[row_offset]
                            cell.value = self._format_employee_name(emp.name)
                            cell.fill = PatternFill(start_color=self.status_colors[status],
                                                    end_color=self.status_colors[status], fill_type='solid')
                            cell.alignment = Alignment(horizontal='center')

                        # Add right border for all cells except the last column
                        if col < table_end_col:
                            cell.border = right_border

                current_row += max_employees_per_day

                # Add totals row (without borders)
                total_row = current_row

                # Write empty cell with total formatting
                cell = ws.cell(row=total_row, column=1)
                cell.fill = self.total_fill

                # Calculate and write totals for each date
                for col, date in enumerate(all_dates, start=2):
                    count = self._count_unique_employees_on_date(sorted_employees, date)
                    cell = ws.cell(row=total_row, column=col)
                    cell.value = count if count > 0 else ''
                    cell.font = Font(bold=True)
                    cell.fill = self.total_fill
                    cell.alignment = Alignment(horizontal='center')
                    # Add right border for all cells except the last column
                    if col < table_end_col:
                        cell.border = right_border

                # Merge location name cells
                ws.merge_cells(start_row=start_row, start_column=1,
                               end_row=total_row, end_column=1)

                # Apply thick border around the entire location table
                for row in range(table_start_row, total_row + 1):
                    # Left border of first column
                    ws.cell(row=row, column=table_start_col).border = Border(
                        left=Side(style='medium'),
                        right=Side(style='thin'),
                        top=Side(style='medium') if row == table_start_row else None,
                        bottom=Side(style='medium') if row == total_row else None
                    )
                    # Right border of last column
                    ws.cell(row=row, column=table_end_col).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='medium'),
                        top=Side(style='medium') if row == table_start_row else None,
                        bottom=Side(style='medium') if row == total_row else None
                    )
                    # Top border for first row
                    if row == table_start_row:
                        for col in range(table_start_col + 1, table_end_col):
                            ws.cell(row=row, column=col).border = Border(
                                top=Side(style='medium'),
                                right=Side(style='thin')
                            )
                    # Bottom border for last row
                    elif row == total_row:
                        for col in range(table_start_col + 1, table_end_col):
                            ws.cell(row=row, column=col).border = Border(
                                bottom=Side(style='medium'),
                                right=Side(style='thin')
                            )

                current_row += 1

        # Adjust column widths
        for col in range(1, len(all_dates) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Freeze panes - freeze top row and first column
        ws.freeze_panes = 'B2'

    def _count_unique_employees_on_date(self, employees, date):
        """Count unique employees on leave for a given date."""
        return len({emp.emp_code for emp in employees if date in emp.leave_dates})

    def _get_employees_by_date(self, employees, date):
        """Get employees who have leave on a specific date, with their status."""
        result = []
        for emp in employees:
            if date in emp.leave_dates:
                result.append((emp, emp.leave_dates[date].status))
        return result  # Add this return statement

    def _get_all_leave_dates(self):
        """Collect all unique leave dates from all employees."""
        all_dates = set()
        for employee in self.employees.values():
            all_dates.update(employee.leave_dates.keys())
        return sorted(list(all_dates))

    def _format_date_header(self, date):
        """Format date as 'Mon 25/11'."""
        return date.strftime('%a %d/%m')

    def _format_employee_name(self, full_name):
        """Format employee name as 'FirstName L.'"""
        parts = full_name.strip().split()
        if len(parts) > 1:
            return f"{parts[0]} {parts[-1][0]}"
        return full_name

    def _write_date_headers(self, ws, row, all_dates):
        """Write date headers starting from column B."""
        for col, date in enumerate(all_dates, start=2):  # Changed from start=1 to start=2
            cell = ws.cell(row=row, column=col)
            cell.value = self._format_date_header(date)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thick_border

    def _get_location_departments(self, location_employees):
        """Get all departments for a location and their employees, sorted by leave count."""
        dept_employees = {}

        for emp in location_employees.values():
            for work_area in emp.work_areas:
                if work_area.department not in dept_employees:
                    dept_employees[work_area.department] = []
                dept_employees[work_area.department].append(emp)

        # Remove duplicates and sort by leave count within each department
        for dept in dept_employees:
            unique_emps = []
            seen = set()
            for emp in dept_employees[dept]:
                if emp.emp_code not in seen:
                    unique_emps.append(emp)
                    seen.add(emp.emp_code)

            # Sort by leave count (descending) and then by name
            dept_employees[dept] = sorted(unique_emps, key=lambda x: (-len(x.leave_dates), x.name))

        return dict(sorted(dept_employees.items()))

    def _get_employees_for_location(self, location):
        """Get employees for a specific location or combined locations."""
        if isinstance(location, tuple):
            # Handle combined locations
            combined_name, location_list = location
            return {code: emp for code, emp in self.employees.items() if
                    any(area.location in location_list for area in emp.work_areas)}
        else:
            # Handle single location
            return {code: emp for code, emp in self.employees.items() if
                    any(area.location == location for area in emp.work_areas)}

    def _get_all_locations(self):
        """Get all unique locations including combined ones."""
        locations = set()
        for employee in self.employees.values():
            for work_area in employee.work_areas:
                locations.add(work_area.location)
        return sorted(list(locations))

    def save_report(self, directory):
        """Save the report with the specified naming convention."""
        current_date = datetime.now()
        month_year = current_date.strftime("%d %b %Y")
        filename = f"Departmental Leave Report {month_year}.xlsx"
        filepath = os.path.join(directory, filename)
        self.wb.save(filepath)
        print(f"Saved: {os.path.basename(filename)}")


def generate_departmental_leave_report(employees, output_directory, employee_manager):
    """Main function to generate and save the departmental leave report."""
    report_generator = DepartmentalLeaveReportGenerator(employees, employee_manager)
    report_generator.generate_report()
    report_generator.save_report(output_directory)
