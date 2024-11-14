import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class LeaveReportGenerator:
    def __init__(self, employees):
        self.employees = {code: emp for code, emp in employees.items() if len(emp.leave_dates) > 0}
        self.wb = Workbook()
        self.ws = None
        self.colors = ['FF0000', '00FF00', '0000FF', 'FFFF00', '00FFFF', 'FF00FF']
        self.status_colors = {}
        self.COMBINED_LOCATIONS = [
            ("ADELAIDE HILLS & STRATHALBYN", ["ADELAIDE HILLS", "STRATHALBYN"]),
        ]

    def _get_all_leave_dates(self):
        """Collect all unique leave dates from all employees."""
        all_dates = set()
        for employee in self.employees.values():
            all_dates.update(employee.leave_dates.keys())
        return sorted(list(all_dates))

    def _format_date_header(self, date):
        """Format date as 'Mon 25/11'."""
        return date.strftime('%a %d/%m')

    def _get_status_initial(self, status):
        """Get first character of status."""
        return status[0] if status else ''

    def _assign_status_colors(self):
        """Assign colors to unique statuses."""
        unique_statuses = set()
        for employee in self.employees.values():
            for leave_date in employee.leave_dates.values():
                unique_statuses.add(leave_date.status)

        for i, status in enumerate(sorted(unique_statuses)):
            color = self.colors[i % len(self.colors)]
            self.status_colors[status] = color

    def _get_all_locations(self):
        """Get all unique locations including combined ones."""
        locations = set()
        for employee in self.employees.values():
            for work_area in employee.work_areas:
                locations.add(work_area.location)
        return sorted(list(locations))

    def _get_employees_for_location(self, location, include_combined=False):
        """Get employees for a specific location or combined locations."""
        if isinstance(location, tuple):
            # Handle combined locations
            combined_name, location_list = location
            return {
                code: emp for code, emp in self.employees.items()
                if any(area.location in location_list for area in emp.work_areas)
            }
        else:
            # Handle single location
            return {
                code: emp for code, emp in self.employees.items()
                if any(area.location == location for area in emp.work_areas)
            }

    def _calculate_status_totals(self, filtered_employees, all_dates):
        """Calculate totals per status and overall total for each date."""
        status_totals = {}
        overall_totals = {}

        # Initialize totals dictionaries
        for date in all_dates:
            status_totals[date] = {}
            overall_totals[date] = 0

        # Calculate totals
        for employee in filtered_employees.values():
            for date in all_dates:
                if date in employee.leave_dates:
                    status = employee.leave_dates[date].status
                    if status not in status_totals[date]:
                        status_totals[date][status] = 0
                    status_totals[date][status] += 1
                    overall_totals[date] += 1

        return status_totals, overall_totals

    def _generate_worksheet(self, ws_name, filtered_employees, all_dates):
        """Generate a worksheet for the given employees."""
        ws = self.wb.create_sheet(ws_name)

        # Define border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        # Set up headers
        headers = ['Employee Name (Code)', 'Employment Type', 'Leave Count']
        headers.extend([self._format_date_header(date) for date in all_dates])

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thick_border

        # Write employee data
        row = 2
        for emp_code, employee in sorted(filtered_employees.items(), key=lambda x: x[1].name):
            # Employee name and code
            cell = ws.cell(row=row, column=1, value=f"{employee.name} ({emp_code})")
            cell.border = thin_border

            # Employment type
            cell = ws.cell(row=row, column=2, value=employee.employment_type)
            cell.border = thin_border

            # Leave count
            cell = ws.cell(row=row, column=3, value=len(employee.leave_dates))
            cell.border = thin_border

            # Status for each date
            for col, date in enumerate(all_dates, 4):
                cell = ws.cell(row=row, column=col)
                if date in employee.leave_dates:
                    leave_date = employee.leave_dates[date]
                    status = self._get_status_initial(leave_date.status)
                    cell.value = status
                    cell.fill = PatternFill(start_color=self.status_colors[leave_date.status],
                                            end_color=self.status_colors[leave_date.status],
                                            fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            row += 1

        # Calculate totals
        status_totals, overall_totals = self._calculate_status_totals(filtered_employees, all_dates)

        # Add spacing row
        row += 1

        # Write status totals
        for status in sorted(self.status_colors.keys()):
            # Status row header
            cell = ws.cell(row=row, column=1, value=f"Total {status}")
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = PatternFill(start_color=self.status_colors[status],
                                    end_color=self.status_colors[status],
                                    fill_type='solid')

            # Empty employment type cell
            ws.cell(row=row, column=2).border = thin_border

            # Calculate total leave count for this status
            status_leave_count = sum(totals.get(status, 0) for totals in status_totals.values())
            cell = ws.cell(row=row, column=3, value=status_leave_count)
            cell.border = thin_border
            cell.font = Font(bold=True)

            # Write daily totals for this status
            for col, date in enumerate(all_dates, 4):
                cell = ws.cell(row=row, column=col)
                count = status_totals[date].get(status, 0)
                if count > 0:
                    cell.value = count
                    cell.fill = PatternFill(start_color=self.status_colors[status],
                                            end_color=self.status_colors[status],
                                            fill_type='solid')
                cell.border = thin_border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            row += 1

        # Write overall totals
        cell = ws.cell(row=row, column=1, value="TOTAL ALL LEAVE")
        cell.font = Font(bold=True)
        cell.border = thick_border
        cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

        # Empty employment type cell
        ws.cell(row=row, column=2).border = thick_border

        # Calculate total leave count across all statuses
        total_leave_count = sum(overall_totals.values())
        cell = ws.cell(row=row, column=3, value=total_leave_count)
        cell.border = thick_border
        cell.font = Font(bold=True)

        # Write daily totals
        for col, date in enumerate(all_dates, 4):
            cell = ws.cell(row=row, column=col)
            count = overall_totals[date]
            if count > 0:
                cell.value = count
                cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            cell.border = thick_border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        row += 2

        # Add legend
        ws.cell(row=row, column=1, value="Legend:").font = Font(bold=True)
        row += 1
        for status, color in self.status_colors.items():
            cell = ws.cell(row=row, column=1)
            cell.value = f"{status} ({status[0]})"
            cell.alignment = Alignment(horizontal='left')

            color_cell = ws.cell(row=row, column=2)
            color_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            row += 1

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            if col <= 2:
                ws.column_dimensions[column_letter].width = 30
            elif col == 3:
                ws.column_dimensions[column_letter].width = 12
            else:
                ws.column_dimensions[column_letter].width = 10

    def generate_report(self):
        """Generate the leave report in Excel format."""
        # Assign colors to statuses
        self._assign_status_colors()

        # Get all unique dates
        all_dates = self._get_all_leave_dates()

        # Remove default sheet if it exists
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])

        # Generate GLOBAL worksheet
        self._generate_worksheet('GLOBAL', self.employees, all_dates)

        # Generate location-specific worksheets
        locations = self._get_all_locations()
        for location in locations:
            filtered_employees = self._get_employees_for_location(location)
            if filtered_employees:  # Only create sheet if there are employees
                safe_name = str(location)[:31]  # Excel worksheet names limited to 31 chars
                self._generate_worksheet(safe_name, filtered_employees, all_dates)

        # Generate combined location worksheets
        for combined_name, location_list in self.COMBINED_LOCATIONS:
            filtered_employees = self._get_employees_for_location((combined_name, location_list))
            if filtered_employees:
                safe_name = combined_name[:31]
                self._generate_worksheet(safe_name, filtered_employees, all_dates)

    def save_report(self, directory):
        """Save the report with the specified naming convention."""
        current_date = datetime.now()
        month_year = current_date.strftime("%b %Y").title()  # Will output like "Aug 2024"
        filename = f"Leave Report {month_year}.xlsx"
        filepath = os.path.join(directory, filename)
        self.wb.save(filepath)
        print(f"Report saved: {filepath}")


def generate_leave_report(employees, output_directory):
    """Main function to generate and save the leave report."""
    report_generator = LeaveReportGenerator(employees)
    report_generator.generate_report()
    report_generator.save_report(output_directory)